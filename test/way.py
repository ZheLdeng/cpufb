import subprocess
from openpyxl import Workbook

file_path = "new_way.cpp"
new_line_prefix = "#define WINDOW_SIZE "
values = [0.015625,0.03125,0.0625,0.125,0.25,0.5,1]
# 2, 4, 8, 16, 32, 64, 128, 256, 512, 1024, 2048, 4096, 8192
n=2048
# 创建 Excel 工作簿和工作表
wb = Workbook()
ws = wb.active

# 写入列标题
ws.append(["Value"] + list(range(1, n+1)))  # 列标题包括 Value 和 1 到 32

# 遍历每个值
for value in values:
    new_line = f"{new_line_prefix}{int(value*1024)}"
    
    # 修改文件内容
    with open(file_path, 'r') as f:
        lines = f.readlines()
        lines[14] = new_line + '\n'
    with open(file_path, 'w') as f:
        f.writelines(lines)
    
    # 编译并运行程序
    subprocess.run(["g++","-O0", "-o", "way", "new_way.cpp"], check=True)
    output = subprocess.run(["./way"], capture_output=True, text=True).stdout
    
    # 将输出按行分割并写入 Excel 表格中的对应列（以数值形式）
    output_lines = output.strip().split('\n')[:n]  # 只取前 32 行输出
    col_values = [value] + [float(line) for line in output_lines]
    ws.append(col_values)

# 转置数据
transposed_data = [[ws.cell(row=i, column=j).value for i in range(1, len(values) + 2)] for j in range(1, n+1)]

# 清空表格
ws.delete_rows(1, ws.max_row)

# 写入转置后的数据
for row in transposed_data:
    ws.append(row)

# 保存 Excel 文件
wb.save("output.xlsx")
